# -*- coding: utf-8 -*-

import json
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from settings import HEADERS


def write_request_sheet(ws, request_meta: Dict[str, Any]):
    ws.title = "Запрос"
    ws.append(["Параметр", "Значение"])

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for k, v in request_meta.items():
        if isinstance(v, (dict, list)):
            ws.append([k, json.dumps(v, ensure_ascii=False)])
        else:
            ws.append([k, str(v)])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_companies_sheet(ws, companies: List[Dict[str, Any]]):
    ws.title = "Организации"
    ws.append(HEADERS)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in companies:
        ws.append([c.get(h, "") for h in HEADERS])

    raw_col_idx = None
    for i, h in enumerate(HEADERS, start=1):
        if h == "raw_json":
            raw_col_idx = i
            break

    align_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_nowrap = Alignment(horizontal="left", vertical="top", wrap_text=False)

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if raw_col_idx is not None and c == raw_col_idx:
                cell.alignment = align_nowrap
            else:
                cell.alignment = align_wrap

    for colnum, header in enumerate(HEADERS, start=1):
        colletter = get_column_letter(colnum)
        maxlen = max(10, len(header) + 2)
        for row in ws.iter_rows(min_row=1, min_col=colnum, max_col=colnum, max_row=ws.max_row):
            cell = row[0]
            if cell.value is None:
                continue
            s = str(cell.value)
            maxlen = max(maxlen, min(len(s), 60))
        ws.column_dimensions[colletter].width = min(max(maxlen, 10), 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def save_to_excel(companies: List[dict], out_path: str, request_meta: Dict[str, Any]):
    wb = openpyxl.Workbook()
    ws_org = wb.active
    write_companies_sheet(ws_org, companies)
    ws_req = wb.create_sheet()
    write_request_sheet(ws_req, request_meta)
    wb.active = 0
    wb.save(out_path)
