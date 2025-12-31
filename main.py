#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Yandex Maps (API Поиска по организациям) -> Excel (1 файл на запуск).

Требования:
- Все параметры задаются в шапке (константы ниже).
- Никакого ввода с консоли.
- Один xlsx-файл, имя включает дату/время.
- Внутри файла: лист "Организации" (компактно, без дублей) + лист "Запрос" (параметры/время).
- Логи в консоль лаконичные: 1-я строка (файл+параметры), далее короткий прогресс.
"""

import json
import math
import os
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple

import requests
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

# ===================== ЗАГРУЗКА .env =====================

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env", override=True)

# ===================== ШАПКА: МЕНЯЕТЕ ТОЛЬКО ЭТО =====================

API_KEY = os.getenv("YM_API_KEY", "").strip()
if not API_KEY:
    raise RuntimeError("YM_API_KEY не задан. Создайте .env (см. .env.example) и укажите ключ.")

TEXT = "резка металла по размерам ленточнопил"
LANG = "ru_RU"

CENTER_LON = 37.6173
CENTER_LAT = 55.7558

DIAMETER_KM = 40

RESULTS_PER_PAGE = 50
MAX_SKIP = 1000
SLEEP_SEC = 0.25

OUT_DIR = "."
OUT_PREFIX = "out"

LOG_EVERY_PAGE = True

MAX_PHONES = 3
MAX_EMAILS = 3
MAX_FAXES = 3

MAX_CATEGORIES_MAIN = 3

# ===================== ДОП. ОБОГАЩЕНИЕ (ОТКЛЮЧАЕМОЕ) =====================

# 1) Официальный доп. запрос по uri (документированный параметр запроса).
ENABLE_URI_REQUERY = True

# 2) Неофициальное обогащение из web-карточки (может перестать работать в любой момент).
# Если True — всегда делаем web и заполняем рейтинг/отзывы (и debug).
ENABLE_UNOFFICIAL_ENRICH = False

# 3) ВАЖНО: fallback для рейтинга.
# Если включен, то даже при ENABLE_UNOFFICIAL_ENRICH=False:
# когда uri-requery не дал рейтинг/отзывы, скрипт автоматически доберет их с web-карточки.
ENABLE_WEB_FALLBACK_FOR_RATING = True

# Детализированный debug-вывод/дампы (для проверки, что реально приходит)
EXTRA_DEBUG = False

EXTRA_MAX_ITEMS = 30
EXTRA_DEBUG_DIR = "debug_extra"

EXTRA_SAVE_URI_JSON = True
EXTRA_SAVE_URI_MATCHES = True

EXTRA_SAVE_WEB_EXTRACT_JSON = True
EXTRA_SAVE_WEB_HTML = False

# ===================== ВСПОМОГАТЕЛЬНОЕ =====================

def now_str_for_filename() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def now_iso_local() -> str:
    return datetime.now().astimezone().replace(microsecond=0).isoformat()


def bbox_from_center_diameter_km(center_lon: float, center_lat: float, diameter_km: float) -> str:
    if diameter_km <= 0:
        raise ValueError("DIAMETER_KM должен быть > 0")

    radius_km = diameter_km / 2.0
    km_per_deg_lat = 110.574
    km_per_deg_lon = 111.320 * math.cos(math.radians(center_lat))
    if abs(km_per_deg_lon) < 1e-9:
        km_per_deg_lon = 1e-9

    dlat = radius_km / km_per_deg_lat
    dlon = radius_km / km_per_deg_lon

    lon1 = center_lon - dlon
    lon2 = center_lon + dlon
    lat1 = center_lat - dlat
    lat2 = center_lat + dlat

    lon1 = max(-180.0, min(180.0, lon1))
    lon2 = max(-180.0, min(180.0, lon2))
    lat1 = max(-90.0, min(90.0, lat1))
    lat2 = max(-90.0, min(90.0, lat2))

    return f"{lon1:.6f},{lat1:.6f}~{lon2:.6f},{lat2:.6f}"


def safe_str(x: Any) -> str:
    if x is None:
        return ""
    if isinstance(x, str):
        return x.strip()
    return str(x).strip()


def dedup_keep_order(items: List[str]) -> List[str]:
    seen = set()
    out = []
    for x in items:
        x = safe_str(x)
        if not x:
            continue
        if x in seen:
            continue
        out.append(x)
        seen.add(x)
    return out


def safe_join(items: List[str]) -> str:
    return "; ".join(dedup_keep_order(items))


def pick_n(items: List[str], n: int) -> List[str]:
    items = items[:n]
    if len(items) < n:
        items = items + [""] * (n - len(items))
    return items


def normalize_hhmm(t: str) -> str:
    t = safe_str(t)
    if len(t) >= 5 and t[2] == ":":
        return t[:5]
    return t


def days_ranges_ru(days: List[str]) -> str:
    order = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"]
    idx = [order.index(d) for d in days if d in order]
    idx = sorted(set(idx))
    if not idx:
        return ""

    ranges: List[Tuple[int, int]] = []
    start = idx[0]
    prev = idx[0]
    for i in idx[1:]:
        if i == prev + 1:
            prev = i
            continue
        ranges.append((start, prev))
        start = i
        prev = i
    ranges.append((start, prev))

    parts = []
    for a, b in ranges:
        if a == b:
            parts.append(order[a])
        else:
            parts.append(f"{order[a]}-{order[b]}")
    return ", ".join(parts)


def bool_to_ru(v: Any) -> str:
    if v is True:
        return "Есть"
    if v is False:
        return "Нет"
    return safe_str(v)


def ensure_debug_dir() -> Path:
    d = BASE_DIR / EXTRA_DEBUG_DIR
    d.mkdir(parents=True, exist_ok=True)
    return d


def dump_debug_json(filename: str, data: Any):
    if not EXTRA_DEBUG:
        return
    d = ensure_debug_dir()
    p = d / filename
    try:
        p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


def dump_debug_text(filename: str, text: str):
    if not EXTRA_DEBUG:
        return
    d = ensure_debug_dir()
    p = d / filename
    try:
        p.write_text(text or "", encoding="utf-8")
    except Exception:
        pass


def short_keys(dct: Any, limit: int = 60) -> str:
    if not isinstance(dct, dict):
        return ""
    keys = list(dct.keys())
    if len(keys) > limit:
        keys = keys[:limit] + ["..."]
    return ", ".join(map(str, keys))


def oid_from_uri(uri: str) -> str:
    m = re.search(r"[?&]oid=(\d+)", safe_str(uri))
    return m.group(1) if m else ""


def find_interest_fields(obj: Any) -> List[Dict[str, Any]]:
    """
    Диагностика: рекурсивно ищет любые ключи, где в названии встречается rating/review.
    """
    out: List[Dict[str, Any]] = []
    patterns = ("rating", "review")

    def walk(x: Any, path: List[str]):
        if isinstance(x, dict):
            for k, v in x.items():
                k_str = safe_str(k)
                new_path = path + [k_str]
                lk = k_str.lower()
                if any(p in lk for p in patterns):
                    out.append({"path": ".".join(new_path), "value": v})
                walk(v, new_path)
        elif isinstance(x, list):
            for i, it in enumerate(x):
                walk(it, path + [f"[{i}]"])

    walk(obj, [])
    return out


# ===================== ПАРСИНГ ОТВЕТА =====================

def parse_contacts(meta: Dict[str, Any]) -> Tuple[List[str], List[str], List[str]]:
    phones: List[str] = []
    emails: List[str] = []
    faxes: List[str] = []

    for contact in (meta.get("Phones", []) or []):
        if not isinstance(contact, dict):
            continue
        formatted = safe_str(contact.get("formatted", ""))
        ctype = safe_str(contact.get("type", "")).lower()
        if not formatted:
            continue

        if ctype == "email":
            emails.append(formatted)
        elif ctype == "fax":
            faxes.append(formatted)
        else:
            phones.append(formatted)

    return dedup_keep_order(phones), dedup_keep_order(emails), dedup_keep_order(faxes)


def parse_categories(meta: Dict[str, Any]) -> List[str]:
    cats = meta.get("Categories", []) or []
    names: List[str] = []
    for c in cats:
        if not isinstance(c, dict):
            continue
        n = safe_str(c.get("name", ""))
        if n:
            names.append(n)
    return dedup_keep_order(names)


def parse_address(meta: Dict[str, Any], props: Dict[str, Any]) -> Tuple[str, str]:
    address = safe_str(meta.get("address", ""))
    addr_obj = meta.get("Address") or {}
    postal = ""

    if isinstance(addr_obj, dict):
        postal = safe_str(addr_obj.get("postal_code", ""))
        formatted = safe_str(addr_obj.get("formatted", ""))
        if not address and formatted:
            address = formatted

    if not address:
        address = safe_str(props.get("description", ""))

    return address, postal


def parse_hours(meta: Dict[str, Any]) -> str:
    hours = meta.get("Hours") or {}
    if not isinstance(hours, dict):
        return ""

    hours_text = safe_str(hours.get("text", ""))
    if hours_text:
        return hours_text

    av = hours.get("Availabilities") or []
    if not isinstance(av, list) or not av:
        return ""

    day_map = {
        "Monday": "пн",
        "Tuesday": "вт",
        "Wednesday": "ср",
        "Thursday": "чт",
        "Friday": "пт",
        "Saturday": "сб",
        "Sunday": "вс",
    }

    parts: List[str] = []
    for a in av:
        if not isinstance(a, dict):
            continue

        if a.get("TwentyFourHours") is True:
            parts.append("круглосуточно")
            continue

        intervals = a.get("Intervals") or []
        segs: List[str] = []
        if isinstance(intervals, list):
            for inter in intervals:
                if not isinstance(inter, dict):
                    continue
                fr = normalize_hhmm(safe_str(inter.get("from", "")))
                to = normalize_hhmm(safe_str(inter.get("to", "")))
                if fr and to:
                    segs.append(f"{fr}–{to}")

        time_str = ", ".join(segs) if segs else ""

        if a.get("Everyday") is True:
            parts.append(f"ежедневно {time_str}".strip())
            continue

        days = []
        for k, ru in day_map.items():
            if a.get(k) is True:
                days.append(ru)

        days_str = days_ranges_ru(days)
        if days_str and time_str:
            parts.append(f"{days_str} {time_str}")
        elif days_str:
            parts.append(days_str)
        elif time_str:
            parts.append(time_str)

    return safe_join(parts)


def parse_features(meta: Dict[str, Any]) -> str:
    feats = meta.get("Features") or []
    if not isinstance(feats, list) or not feats:
        return ""

    out: List[str] = []
    for f in feats:
        if not isinstance(f, dict):
            continue

        name = safe_str(f.get("name", "")) or safe_str(f.get("id", ""))
        value = f.get("value")
        value_str = ""

        if isinstance(value, bool):
            value_str = bool_to_ru(value)
        elif isinstance(value, list):
            names = []
            for x in value:
                if isinstance(x, dict):
                    n = safe_str(x.get("name", "")) or safe_str(x.get("id", ""))
                    if n:
                        names.append(n)
                else:
                    s = bool_to_ru(x)
                    if s:
                        names.append(s)
            value_str = safe_join(names)
        elif isinstance(value, dict):
            value_str = safe_str(value.get("name", "")) or safe_str(value.get("id", "")) or safe_str(value)
        else:
            value_str = bool_to_ru(value)

        if name and value_str:
            out.append(f"{name}: {value_str}")
        elif name:
            out.append(name)
        elif value_str:
            out.append(value_str)

    s = safe_join(out)
    s = s.replace(": True", ": Есть").replace(": False", ": Нет")
    return s


def extract_company_info(feature: dict) -> Optional[dict]:
    try:
        props = feature.get("properties", {}) or {}
        meta = props.get("CompanyMetaData", {}) or {}

        coords = (feature.get("geometry", {}) or {}).get("coordinates", []) or []
        lon = coords[0] if len(coords) >= 1 else ""
        lat = coords[1] if len(coords) >= 2 else ""

        org_id = safe_str(meta.get("id", ""))
        name = safe_str(meta.get("name", "")) or safe_str(props.get("name", ""))

        address, postal = parse_address(meta, props)

        phones, emails, faxes = parse_contacts(meta)
        phones_cols = pick_n(phones, MAX_PHONES)
        emails_cols = pick_n(emails, MAX_EMAILS)
        faxes_cols = pick_n(faxes, MAX_FAXES)

        categories = parse_categories(meta)
        cat_main = categories[:MAX_CATEGORIES_MAIN]
        cat_main_cols = pick_n(cat_main, MAX_CATEGORIES_MAIN)
        cat_extra = categories[MAX_CATEGORIES_MAIN:]
        cat_extra_str = safe_join(cat_extra)

        worktime = parse_hours(meta)
        features_str = parse_features(meta)

        rating = meta.get("rating", "")
        review_count = meta.get("review_count", "")

        row: Dict[str, Any] = {
            "ID": org_id,
            "Название": name,
            "Адрес": address,
            "Индекс": postal,
            "Долгота": lon,
            "Широта": lat,
            "Сайт": safe_str(meta.get("url", "")),
            "Телефон 1": phones_cols[0],
            "Телефон 2": phones_cols[1],
            "Телефон 3": phones_cols[2],
            "Email 1": emails_cols[0],
            "Email 2": emails_cols[1],
            "Email 3": emails_cols[2],
            "Режим работы": worktime,
            "Рейтинг": rating,
            "Количество отзывов": review_count,
            "Категория 1": cat_main_cols[0],
            "Категория 2": cat_main_cols[1],
            "Категория 3": cat_main_cols[2],
            "Особенности": features_str,
            "uri": safe_str(props.get("uri", "")),
            "Факс 1": faxes_cols[0],
            "Факс 2": faxes_cols[1],
            "Факс 3": faxes_cols[2],
            "Категории (прочие)": cat_extra_str,
            "raw_json": json.dumps(feature, ensure_ascii=False),
        }
        return row
    except Exception:
        return None


# ===================== API =====================

def ymaps_get_json(session: requests.Session, params: Dict[str, Any]) -> Dict[str, Any]:
    url = "https://search-maps.yandex.ru/v1/"

    retry_statuses = {429, 500, 502, 503, 504}
    max_retries = 6
    backoff = 1.0
    last_err = None

    for _ in range(max_retries):
        try:
            r = session.get(url, params=params, timeout=25)

            if r.status_code in retry_statuses:
                body = (r.text or "").strip().replace("\n", " ")
                last_err = f"{r.status_code} {r.reason}: {body[:300]}"
                time.sleep(backoff)
                backoff *= 2
                continue

            if r.status_code >= 400:
                body = (r.text or "").strip().replace("\n", " ")
                raise requests.HTTPError(f"{r.status_code} {r.reason}: {body[:2000]}", response=r)

            return r.json()

        except (requests.Timeout, requests.ConnectionError) as e:
            last_err = str(e)
            time.sleep(backoff)
            backoff *= 2

    raise requests.HTTPError(f"retry_failed: {last_err}")


def ymaps_search_page(session: requests.Session, bbox: str, skip: int) -> List[dict]:
    params = {
        "apikey": API_KEY,
        "text": TEXT,
        "lang": LANG,
        "type": "biz",
        "bbox": bbox,
        "rspn": 1,
        "results": RESULTS_PER_PAGE,
        "skip": skip,
    }

    data = ymaps_get_json(session, params=params)
    features = data.get("features", []) or []

    rows = []
    for f in features:
        row = extract_company_info(f)
        if row:
            rows.append(row)
    return rows


def ymaps_fetch_by_uri(session: requests.Session, uri: str) -> Dict[str, Any]:
    params = {
        "apikey": API_KEY,
        "uri": safe_str(uri),
        "lang": LANG,
        "type": "biz",
        "results": 1,
        "skip": 0,
    }
    return ymaps_get_json(session, params=params)


def web_fetch_org_page(session: requests.Session, oid: str) -> str:
    url = f"https://yandex.ru/maps/org/{oid}/"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/120.0 Safari/537.36",
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
    }
    r = session.get(url, headers=headers, timeout=25)
    return r.text or ""


def extract_jsonld_blocks(html: str) -> List[Any]:
    blocks = re.findall(
        r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
        html or "",
        flags=re.DOTALL | re.IGNORECASE,
    )
    out = []
    for b in blocks:
        b = (b or "").strip()
        if not b:
            continue
        try:
            out.append(json.loads(b))
        except Exception:
            continue
    return out


def walk_find(obj: Any, key: str) -> List[Any]:
    found = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k == key:
                found.append(v)
            found.extend(walk_find(v, key))
    elif isinstance(obj, list):
        for it in obj:
            found.extend(walk_find(it, key))
    return found


def parse_web_jsonld(jsonlds: List[Any]) -> Dict[str, Any]:
    rating_value = ""
    rating_count = ""
    review_count = ""
    same_as: List[str] = []

    for o in jsonlds:
        for agg in walk_find(o, "aggregateRating"):
            if isinstance(agg, dict):
                if rating_value == "" and agg.get("ratingValue") is not None:
                    rating_value = safe_str(agg.get("ratingValue"))
                if rating_count == "" and agg.get("ratingCount") is not None:
                    rating_count = safe_str(agg.get("ratingCount"))
                if review_count == "" and agg.get("reviewCount") is not None:
                    review_count = safe_str(agg.get("reviewCount"))

        for sa in walk_find(o, "sameAs"):
            if isinstance(sa, str):
                same_as.append(sa)
            elif isinstance(sa, list):
                for it in sa:
                    if isinstance(it, str):
                        same_as.append(it)

    same_as = dedup_keep_order([safe_str(x) for x in same_as if safe_str(x)])

    return {
        "ratingValue": rating_value,
        "ratingCount": rating_count,
        "reviewCount": review_count,
        "sameAs": same_as,
    }


def parse_web_microdata(html: str) -> Dict[str, Any]:
    def find_meta(prop: str) -> str:
        m = re.search(
            rf'itemProp=["\']{re.escape(prop)}["\'][^>]*content=["\']([^"\']+)["\']',
            html or "",
            flags=re.IGNORECASE,
        )
        return safe_str(m.group(1)) if m else ""

    return {
        "ratingValue": find_meta("ratingValue"),
        "ratingCount": find_meta("ratingCount"),
        "reviewCount": find_meta("reviewCount"),
    }


def enrich_rating_from_web(session: requests.Session, oid: str) -> Dict[str, str]:
    """
    Возвращает {"rating": "...", "review_count": "...", "rating_count": "..."} (все строки).
    """
    html = web_fetch_org_page(session, oid)

    if EXTRA_SAVE_WEB_HTML:
        dump_debug_text(f"web_{oid}.html", html)

    jsonlds = extract_jsonld_blocks(html)
    parsed_jsonld = parse_web_jsonld(jsonlds)
    parsed_micro = parse_web_microdata(html)

    rating_value = parsed_jsonld.get("ratingValue") or parsed_micro.get("ratingValue") or ""
    review_count = parsed_jsonld.get("reviewCount") or parsed_micro.get("reviewCount") or ""
    rating_count = parsed_jsonld.get("ratingCount") or parsed_micro.get("ratingCount") or ""

    if EXTRA_SAVE_WEB_EXTRACT_JSON:
        dump_debug_json(
            f"web_{oid}_extract.json",
            {
                "jsonld_blocks_count": len(jsonlds),
                "parsed_jsonld": parsed_jsonld,
                "parsed_microdata": parsed_micro,
                "final": {
                    "ratingValue": rating_value,
                    "reviewCount": review_count,
                    "ratingCount": rating_count,
                },
            },
        )

    if EXTRA_DEBUG:
        print(
            f"  web: jsonld_blocks={len(jsonlds)} "
            f"ratingValue={rating_value} reviewCount={review_count} ratingCount={rating_count}"
        )

    return {
        "rating": safe_str(rating_value),
        "review_count": safe_str(review_count),
        "rating_count": safe_str(rating_count),
    }


def fetch_all(bbox: str):
    all_rows: List[dict] = []
    seen_ids = set()
    error_text = ""

    uri_done = 0
    web_done = 0

    with requests.Session() as session:
        skip = 0

        while skip <= MAX_SKIP:
            try:
                rows = ymaps_search_page(session, bbox=bbox, skip=skip)
            except Exception as e:
                error_text = str(e)
                break

            if not rows:
                break

            for row in rows:
                org_id = safe_str(row.get("ID", ""))
                if org_id:
                    if org_id in seen_ids:
                        continue
                    seen_ids.add(org_id)

                uri = safe_str(row.get("uri", ""))
                oid = oid_from_uri(uri) or safe_str(org_id)

                if (ENABLE_URI_REQUERY or ENABLE_UNOFFICIAL_ENRICH or ENABLE_WEB_FALLBACK_FOR_RATING) and EXTRA_DEBUG:
                    print(f"extra: ID={org_id} uri={uri}")

                # --------- 1) Официальный requery по uri (если включен) ---------
                if ENABLE_URI_REQUERY and uri_done < EXTRA_MAX_ITEMS:
                    if uri:
                        try:
                            j2 = ymaps_fetch_by_uri(session, uri)

                            if EXTRA_SAVE_URI_JSON:
                                dump_debug_json(f"uri_{org_id or oid}.json", j2)

                            feats2 = j2.get("features", []) or []
                            f2 = feats2[0] if isinstance(feats2, list) and feats2 else None
                            props2 = (f2.get("properties", {}) or {}) if isinstance(f2, dict) else {}
                            meta2 = (props2.get("CompanyMetaData", {}) or {}) if isinstance(props2, dict) else {}

                            if EXTRA_DEBUG:
                                print(f"  uri-requery: features={len(feats2)} CompanyMetaData keys=[{short_keys(meta2)}]")

                            if EXTRA_SAVE_URI_MATCHES:
                                matches = find_interest_fields(j2)
                                dump_debug_json(f"uri_{org_id or oid}_matches.json",
                                                {"matches_count": len(matches), "matches": matches})
                                if EXTRA_DEBUG:
                                    print(f"  uri-requery: matches(rating/review)={len(matches)} (saved to debug)")

                            # Если вдруг появилось — заполняем существующие колонки
                            if not safe_str(row.get("Рейтинг", "")) and meta2.get("rating") is not None:
                                row["Рейтинг"] = safe_str(meta2.get("rating"))
                            if not safe_str(row.get("Количество отзывов", "")) and meta2.get("review_count") is not None:
                                row["Количество отзывов"] = safe_str(meta2.get("review_count"))

                            uri_done += 1
                            time.sleep(SLEEP_SEC)
                        except Exception as e:
                            if EXTRA_DEBUG:
                                print(f"  uri-requery: ERROR: {e}")

                # --------- 2) Web enrich (если включен явно) ---------
                if ENABLE_UNOFFICIAL_ENRICH and web_done < EXTRA_MAX_ITEMS:
                    if oid.isdigit():
                        try:
                            w = enrich_rating_from_web(session, oid)
                            if not safe_str(row.get("Рейтинг", "")) and w["rating"]:
                                row["Рейтинг"] = w["rating"]
                            if not safe_str(row.get("Количество отзывов", "")) and w["review_count"]:
                                row["Количество отзывов"] = w["review_count"]
                            web_done += 1
                        except Exception as e:
                            if EXTRA_DEBUG:
                                print(f"  web: ERROR: {e}")

                # --------- 3) Fallback: если после uri нет рейтинга/отзывов — добираем с web ---------
                if ENABLE_WEB_FALLBACK_FOR_RATING and not ENABLE_UNOFFICIAL_ENRICH and web_done < EXTRA_MAX_ITEMS:
                    need_rating = not safe_str(row.get("Рейтинг", ""))
                    need_reviews = not safe_str(row.get("Количество отзывов", ""))

                    if (need_rating or need_reviews) and oid.isdigit():
                        try:
                            w = enrich_rating_from_web(session, oid)
                            if need_rating and w["rating"]:
                                row["Рейтинг"] = w["rating"]
                            if need_reviews and w["review_count"]:
                                row["Количество отзывов"] = w["review_count"]
                            web_done += 1
                        except Exception as e:
                            if EXTRA_DEBUG:
                                print(f"  web-fallback: ERROR: {e}")

                all_rows.append(row)

            if LOG_EVERY_PAGE:
                print(f"skip={skip}: +{len(rows)} (total={len(all_rows)})")

            if len(rows) < RESULTS_PER_PAGE:
                break

            skip += RESULTS_PER_PAGE
            time.sleep(SLEEP_SEC)

    return all_rows, error_text


# ===================== EXCEL =====================

HEADERS = [
    "ID",
    "Название",
    "Адрес",
    "Индекс",
    "Долгота",
    "Широта",
    "Сайт",
    "Телефон 1",
    "Телефон 2",
    "Телефон 3",
    "Email 1",
    "Email 2",
    "Email 3",
    "Режим работы",
    "Рейтинг",
    "Количество отзывов",
    "Категория 1",
    "Категория 2",
    "Категория 3",
    "Особенности",
    "uri",
    "Факс 1",
    "Факс 2",
    "Факс 3",
    "Категории (прочие)",
    "raw_json",
]


def write_request_sheet(ws, request_meta: Dict[str, Any]):
    ws.title = "Запрос"
    ws.append(["Параметр", "Значение"])

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for k, v in request_meta.items():
        ws.append([k, str(v)])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_companies_sheet(ws, companies: List[dict]):
    ws.title = "Организации"
    ws.append(HEADERS)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if companies:
        for c in companies:
            ws.append([c.get(h, "") for h in HEADERS])

    raw_col_idx = None
    for i, h in enumerate(HEADERS, start=1):
        if h == "raw_json":
            raw_col_idx = i
            break

    align_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_no_wrap = Alignment(horizontal="left", vertical="top", wrap_text=False)

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if raw_col_idx is not None and c == raw_col_idx:
                cell.alignment = align_no_wrap
            else:
                cell.alignment = align_wrap

    for col_num, header in enumerate(HEADERS, 1):
        col_letter = get_column_letter(col_num)
        max_len = len(header) + 2

        for row in ws.iter_rows(min_row=1, min_col=col_num, max_col=col_num, max_row=ws.max_row):
            cell = row[0]
            if cell.value is None:
                continue

            s = str(cell.value)
            if header == "raw_json":
                max_len = max(max_len, min(len(s), 60))
            else:
                for part in s.split("\n"):
                    max_len = max(max_len, len(part) + 1)

        ws.column_dimensions[col_letter].width = min(max(max_len, 10), 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def save_to_excel(companies: List[dict], out_path: str, request_meta: Dict[str, Any]):
    wb = openpyxl.Workbook()

    ws_org = wb.active
    write_companies_sheet(ws_org, companies)

    ws_req = wb.create_sheet("Запрос")
    write_request_sheet(ws_req, request_meta)

    wb.active = 0
    wb.save(out_path)


# ===================== MAIN =====================

def main():
    request_time = now_iso_local()
    bbox = bbox_from_center_diameter_km(CENTER_LON, CENTER_LAT, DIAMETER_KM)
    out_name = f"{OUT_PREFIX}_{now_str_for_filename()}.xlsx"
    out_path = os.path.join(OUT_DIR, out_name)

    print(f"{out_name} | text='{TEXT}' | center={CENTER_LON},{CENTER_LAT} | diameter_km={DIAMETER_KM} | lang={LANG}")

    request_meta = {
        "request_time": request_time,
        "text": TEXT,
        "lang": LANG,
        "center_lon": CENTER_LON,
        "center_lat": CENTER_LAT,
        "diameter_km": DIAMETER_KM,
        "bbox": bbox,
        "results_per_page": RESULTS_PER_PAGE,
        "max_skip": MAX_SKIP,
        "sleep_sec": SLEEP_SEC,
        "max_phones": MAX_PHONES,
        "max_emails": MAX_EMAILS,
        "max_faxes": MAX_FAXES,
        "max_categories_main": MAX_CATEGORIES_MAIN,
        "ENABLE_URI_REQUERY": ENABLE_URI_REQUERY,
        "ENABLE_UNOFFICIAL_ENRICH": ENABLE_UNOFFICIAL_ENRICH,
        "ENABLE_WEB_FALLBACK_FOR_RATING": ENABLE_WEB_FALLBACK_FOR_RATING,
    }

    companies, err = fetch_all(bbox=bbox)

    if err:
        print(f"ERROR: {err}")
        request_meta["error"] = err

    print(f"done: rows={len(companies)}")
    save_to_excel(companies, out_path, request_meta)
    print("saved")


if __name__ == "__main__":
    main()
