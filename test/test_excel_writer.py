import openpyxl

from ymaps_excel_export.excel_writer import save_to_excel
from ymaps_excel_export.models import Company


def test_save_to_excel_creates_workbook(st_base, tmp_path):
    out = tmp_path / "t.xlsx"

    companies = [
        Company(
            ID="1",
            Название="Тест",
            Рейтинг="4,9",
            Количество_отзывов="10",
            raw_json="{}",
        )
    ]

    meta = {"hello": "world", "n": 1}
    save_to_excel(st_base, companies, str(out), meta)

    wb = openpyxl.load_workbook(out)
    assert "Организации" in wb.sheetnames
    assert "Запрос" in wb.sheetnames

    ws_org = wb["Организации"]
    headers = [c.value for c in ws_org[1]]
    assert headers == st_base.HEADERS

    ws_req = wb["Запрос"]
    # Заголовок на листе запроса
    assert ws_req["A1"].value == "Параметр"
    assert ws_req["B1"].value == "Значение"
