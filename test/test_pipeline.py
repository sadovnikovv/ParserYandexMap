import openpyxl

from ymaps_excel_export.models import Company, RunResult
from ymaps_excel_export.pipeline import run


def test_pipeline_onlineapi_success_writes_excel(st_base, monkeypatch, tmp_path):
    # Фиксируем имя файла
    import ymaps_excel_export.pipeline as pipe
    monkeypatch.setattr(pipe, "now_str_for_filename", lambda: "TESTTIME")

    # Мокаем API поиск
    def fake_search_bbox(st, bbox):
        return [Company(ID="1", Название="Org", raw_json="{}")], {"total": 1, "unique": 1}, ""

    monkeypatch.setattr(pipe, "search_bbox", fake_search_bbox)

    # Отключаем web-enrich, чтобы тест был стабильнее/быстрее
    st = st_base.__class__(**{**st_base.__dict__, "OFFLINE_ENRICH_MODE": "NONE", "MODE": "ONLINEAPI"})

    res = run(st)
    assert isinstance(res, RunResult)
    assert res.request_meta["rows"] == 1
    assert res.request_meta["saved"].endswith("_TESTTIME.xlsx")

    wb = openpyxl.load_workbook(res.request_meta["saved"])
    assert "Организации" in wb.sheetnames
    ws = wb["Организации"]
    assert ws.max_row == 2  # 1 заголовок + 1 организация


def test_pipeline_onlineapi_error_sets_error_and_rows0(st_base, monkeypatch):
    import ymaps_excel_export.pipeline as pipe
    monkeypatch.setattr(pipe, "now_str_for_filename", lambda: "TESTTIME")

    def fake_search_bbox(st, bbox):
        return [], {"total": 0, "unique": 0}, "Yandex API error: HTTP 403 Forbidden: Invalid apikey. Hint: ..."

    monkeypatch.setattr(pipe, "search_bbox", fake_search_bbox)

    st = st_base.__class__(**{**st_base.__dict__, "OFFLINE_ENRICH_MODE": "NONE", "MODE": "ONLINEAPI"})

    res = run(st)
    assert res.request_meta.get("error")
    assert res.request_meta["rows"] == 0
    assert res.request_meta.get("saved")
