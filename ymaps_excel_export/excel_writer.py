# -*- coding: utf-8 -*-

from __future__ import annotations

import json
import math
import re
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import Settings
from .models import Company

_INT_RE = re.compile(r"^\s*\d+\s*$")


def _to_int_maybe(x: Any) -> Optional[int]:
    if x is None:
        return None
    if isinstance(x, int):
        return x
    s = str(x).strip()
    if not s:
        return None
    s2 = s.replace(" ", "").replace("\u00A0", "")
    if _INT_RE.match(s2):
        try:
            return int(s2)
        except Exception:
            return None
    return None


def _to_float_ru_maybe(x: Any) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, float):
        return x
    if isinstance(x, int):
        return float(x)
    s = str(x).strip()
    if not s:
        return None
    s = s.replace("\u00A0", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _find_col_idx(headers: List[str], name: str) -> Optional[int]:
    try:
        return headers.index(name) + 1  # 1-based
    except ValueError:
        return None


def _cell_lines_estimate(value: Any, col_width_chars: float) -> int:
    """
    Очень грубая оценка, сколько строк займёт ячейка при wrap_text=True.
    Используем ширину колонки как "примерное число символов в строке".
    """
    if value is None:
        return 1
    s = str(value)
    if not s:
        return 1

    cpl = max(int(col_width_chars), 1)  # chars per line
    # Учитываем явные переносы строк
    parts = s.splitlines() or [s]

    total = 0
    for part in parts:
        part = part or ""
        total += max(1, int(math.ceil(len(part) / cpl)))

    return max(1, total)


def write_request_sheet(ws, request_meta: Dict[str, Any]) -> None:
    ws.title = "Запрос"
    ws.append(["Параметр", "Значение"])

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_font = Font(name="Calibri", size=11)
    body_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for k, v in request_meta.items():
        if isinstance(v, (dict, list)):
            ws.append([k, json.dumps(v, ensure_ascii=False)])
        else:
            ws.append([k, str(v)])

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font
            cell.alignment = body_alignment

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_companies_sheet(ws, st: Settings, companies: List[Company]) -> None:
    ws.title = "Организации"
    ws.append(st.HEADERS)

    # Стили
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    body_font = Font(name="Calibri", size=11)
    align_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_nowrap = Alignment(horizontal="left", vertical="top", wrap_text=False)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Данные
    rows = [c.as_excel_row() for c in companies]
    for r in rows:
        ws.append([r.get(h, "") for h in st.HEADERS])

    idx_raw = _find_col_idx(st.HEADERS, "raw_json")
    idx_id = _find_col_idx(st.HEADERS, "ID")
    idx_rating = _find_col_idx(st.HEADERS, "Рейтинг")
    idx_reviews = _find_col_idx(st.HEADERS, "Количество отзывов")

    # 1) Выравнивание + шрифт + числовые форматы
    for rr in range(2, ws.max_row + 1):
        for cc in range(1, ws.max_column + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.font = body_font
            cell.alignment = align_nowrap if (idx_raw is not None and cc == idx_raw) else align_wrap

        # ID — целое (без E+11)
        if idx_id is not None:
            c = ws.cell(row=rr, column=idx_id)
            v = _to_int_maybe(c.value)
            if v is not None:
                c.value = v
                c.number_format = "0"

        # Рейтинг — по центру
        if idx_rating is not None:
            c = ws.cell(row=rr, column=idx_rating)
            v = _to_float_ru_maybe(c.value)
            if v is not None:
                c.value = v
                c.number_format = "0.0"
            c.alignment = align_center

        # Количество отзывов — по центру
        if idx_reviews is not None:
            c = ws.cell(row=rr, column=idx_reviews)
            v = _to_int_maybe(c.value)
            if v is not None:
                c.value = v
                c.number_format = "0"
            c.alignment = align_center

    # 2) Автоподбор ширины колонок (как в твоём “лучшем” варианте)
    for colnum, header in enumerate(st.HEADERS, start=1):
        colletter = get_column_letter(colnum)
        maxlen = max(10, len(header) + 2)

        for row in ws.iter_rows(min_row=1, min_col=colnum, max_col=colnum, max_row=ws.max_row):
            cell = row[0]
            if cell.value is None:
                continue
            s = str(cell.value)
            maxlen = max(maxlen, min(len(s), 60))

        ws.column_dimensions[colletter].width = min(max(maxlen, 10), 60)

    # 3) Высота строк: только 2 варианта (1 строка или 2 строки), Calibri 11
    # В Excel дефолт примерно 15 pt для Calibri 11, значит 2 строки ~= 30 pt.
    ONE_LINE_PT = 15.0
    TWO_LINES_PT = 30.0

    # Для оценки высоты используем wrap-колонки; raw_json (nowrap) не должен влиять на высоту.
    wrap_cols: List[int] = []
    for colnum, header in enumerate(st.HEADERS, start=1):
        if idx_raw is not None and colnum == idx_raw:
            continue
        wrap_cols.append(colnum)

    # Считаем высоту по строкам: если где-то нужно >1 строки — ставим 30, иначе 15.
    for rr in range(1, ws.max_row + 1):
        need_two = False
        for cc in wrap_cols:
            colletter = get_column_letter(cc)
            w = ws.column_dimensions[colletter].width
            w = float(w) if w is not None else 10.0

            v = ws.cell(row=rr, column=cc).value
            if _cell_lines_estimate(v, w) >= 2:
                need_two = True
                break

        ws.row_dimensions[rr].height = TWO_LINES_PT if need_two else ONE_LINE_PT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def save_to_excel(st: Settings, companies: List[Company], out_path: str, request_meta: Dict[str, Any]) -> None:
    wb = openpyxl.Workbook()

    ws_org = wb.active
    write_companies_sheet(ws_org, st, companies)

    ws_req = wb.create_sheet()
    write_request_sheet(ws_req, request_meta)

    wb.active = 0
    wb.save(out_path)
